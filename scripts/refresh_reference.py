#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""knowledge/reference/ 원본(xlsx/xls)에서 파생 데이터(JSON·MD)를 재생성한다.

LLM 개입 없는 결정적 변환. 새 학기 시간표나 갱신된 사업관리카드를 받으면
knowledge/reference/의 원본 파일을 교체한 뒤 이 스크립트만 다시 실행한다.

생성물:
  - 기초시간표-<이름>.json / .md   (주제선택 표기 (주)·㈜·(원) 제거)
  - 예산과목-<연도>.json           (세부사업 > 세부항목 > 비목·산출내역 이름 목록)

사용: python scripts/refresh_reference.py
의존: openpyxl(xlsx), xlrd(구형 xls)
"""
import datetime
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
REF = ROOT / "knowledge" / "reference"

TIMETABLE_XLSX = REF / "기초시간표-2026-2학기.xlsx"
TIMETABLE_STEM = "기초시간표-2026-2학기"
BUDGET_XLS = REF / "사업관리카드-예산-2026.xls"
BUDGET_JSON = REF / "예산과목-2026.json"

DAYS = ["월", "화", "수", "목", "금"]
# 요일별 (1학년 열, 2학년 열) — openpyxl 1-based column index
DAY_COLS = {"월": (2, 3), "화": (4, 5), "수": (6, 7), "목": (8, 9), "금": (10, 11)}
PERIOD_ROWS = {p: 4 + p for p in range(1, 9)}  # 교시 1~8 = 행 5~12
MARKER_RE = re.compile(r"\((주|원)\)|㈜")  # 주제선택 표기 — 무시(2026. 7. 15. 사용자 확인)

SUBJECT_FULL = {
    "국": "국어", "수": "수학", "사": "사회", "역": "역사", "과": "과학",
    "영": "영어", "체": "체육", "미": "미술", "음": "음악",
    "기가": "기술·가정", "진": "진로", "도덕": "도덕", "정보": "정보",
}
NON_SUBJECT = ["동아리", "스포츠클럽", "교과방과후", "피아노, 스포츠, 기초", "밴드, 기초"]


def _merged_value(ws, row, col):
    """병합 셀이면 병합 범위 좌상단 값을 반환."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col).value
    return ws.cell(row, col).value


def build_timetable():
    import openpyxl
    wb = openpyxl.load_workbook(TIMETABLE_XLSX)
    ws = wb.active
    title = str(ws.cell(1, 1).value).strip()
    table = {}
    for day in DAYS:
        c1, c2 = DAY_COLS[day]
        table[day] = {"1": [], "2": []}
        for p in range(1, 9):
            r = PERIOD_ROWS[p]
            for grade, col in (("1", c1), ("2", c2)):
                v = _merged_value(ws, r, col)
                v = "" if v is None else MARKER_RE.sub("", str(v).strip()).strip()
                table[day][grade].append(v)
    # 검증: 모든 칸이 알려진 과목 약어이거나 비교과 항목이어야 함
    unknown = set()
    for day in DAYS:
        for grade in ("1", "2"):
            for v in table[day][grade]:
                if v and v not in SUBJECT_FULL and v not in NON_SUBJECT:
                    unknown.add(v)
    if unknown:
        raise SystemExit(f"[refresh] 알 수 없는 시간표 항목: {sorted(unknown)} — SUBJECT_FULL/NON_SUBJECT 갱신 필요")
    data = {
        "title": title,
        "source": TIMETABLE_XLSX.name,
        "generated": datetime.date.today().isoformat(),
        "grades": {"1": "1학년", "2": "2학년"},
        "subject_full": SUBJECT_FULL,
        "non_subject": NON_SUBJECT,
        "note": "주제선택 표기((주)·㈜·(원))는 원본에 있으나 무시하기로 확정(2026. 7. 15.)하여 제거됨",
        "시간표": table,
    }
    (REF / f"{TIMETABLE_STEM}.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    # 사람용 MD 재생성
    lines = [
        f"# {title}",
        "",
        f"- 원본: `{TIMETABLE_XLSX.name}` (사용자 제공, 2026. 7. 15.) — **이 파일은 `scripts/refresh_reference.py`가 재생성하는 파생 문서. 직접 수정 금지.**",
        "- 각 요일 아래 `1`/`2` 열 = **1학년 / 2학년** (2026. 7. 15. 사용자 확인).",
        "- 원본의 주제선택 표기 `(주)`·`㈜`·`(원)`은 무시하기로 확정되어 제거됨.",
        "- 융합교과 산출은 이 md가 아니라 `scripts/fusion_timetable.py`(기계 계산)를 사용한다. 요일 LLM 암산 금지.",
        "- 한계: 단축수업·고사기간·학사행사 등 특별시간표 미반영 → 활동일이 걸릴 가능성 있으면 사용자 확인.",
        "",
        "| 교시 | " + " | ".join(f"{d}·{g}" for d in DAYS for g in ("1", "2")) + " |",
        "|---|" + "---|" * 10,
    ]
    for p in range(1, 9):
        cells = [table[d][g][p - 1] for d in DAYS for g in ("1", "2")]
        lines.append(f"| {p} | " + " | ".join(cells) + " |")
    lines += [
        "",
        "## 구조 메모",
        "- 월요일만 7~8교시가 정규 교과. 화~금 7~8교시는 방과후 프로그램.",
        "- 금요일 2교시(동아리)·3교시(스포츠클럽)는 1·2학년 공통.",
        "- 과목 약어: " + ", ".join(f"{k}={v}" for k, v in SUBJECT_FULL.items()) + ".",
        "",
    ]
    (REF / f"{TIMETABLE_STEM}.md").write_text("\n".join(lines), encoding="utf-8")
    return data


def build_budget():
    import xlrd
    book = xlrd.open_workbook(str(BUDGET_XLS))
    sheet = book.sheets()[0]
    snapshot = datetime.date.fromtimestamp(BUDGET_XLS.stat().st_mtime).isoformat()
    programs = []
    cur_prog = cur_item = None
    for r in range(sheet.nrows):
        raw = sheet.cell_value(r, 0)
        if not isinstance(raw, str) or not raw.strip():
            continue
        name = raw.strip()
        indent = len(raw) - len(raw.lstrip())
        detail = str(sheet.cell_value(r, 1)).strip() if sheet.ncols > 1 else ""
        amount = sheet.cell_value(r, 2) if sheet.ncols > 2 else ""
        amount = int(amount) if isinstance(amount, float) else None
        if name in ("사업관리카드(예산)",) or "합 계" in name:
            continue
        if r <= 3:  # 헤더 행
            continue
        if detail:  # 원가통계비목 행: col0=비목, col1=산출내역
            if cur_item is None:
                raise SystemExit(f"[refresh] {r}행: 상위 세부항목 없이 산출내역 등장: {name}/{detail}")
            cur_item["산출내역"].append({"비목": name, "명": detail, "예산현액": amount})
        elif indent <= 5:  # 세부사업
            cur_prog = {"세부사업": name, "예산현액": amount, "세부항목": []}
            programs.append(cur_prog)
            cur_item = None
        else:  # 세부항목
            if cur_prog is None:
                raise SystemExit(f"[refresh] {r}행: 상위 세부사업 없이 세부항목 등장: {name}")
            cur_item = {"명": name, "예산현액": amount, "산출내역": []}
            cur_prog["세부항목"].append(cur_item)
    n_names = sum(len(i["산출내역"]) for p in programs for i in p["세부항목"])
    if not programs or n_names == 0:
        raise SystemExit("[refresh] 사업관리카드 파싱 결과가 비었음 — 파일 구조 변경 여부 확인 필요")
    data = {
        "source": BUDGET_XLS.name,
        "snapshot": snapshot,
        "note": "예산과목 '이름' 대조가 목적(2026. 7. 15. 사용자 확인). 금액은 스냅샷 참고치일 뿐 검증 대상 아님.",
        "사업": programs,
    }
    BUDGET_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return len(programs), n_names


def main():
    tt = build_timetable()
    n_days = len(tt["시간표"])
    n_prog, n_names = build_budget()
    print(f"[refresh] 시간표 JSON/MD 재생성 완료: {n_days}개 요일 x 2학년 x 8교시")
    print(f"[refresh] 예산과목 JSON 재생성 완료: 세부사업 {n_prog}개, 산출내역 {n_names}개")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
